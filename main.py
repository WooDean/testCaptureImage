import streamlit as st
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.utils import coordinate_to_tuple
import tempfile
from PIL import Image as PILImage

# 스타일 설정
st.markdown(
    """
    <style>
    body {
        color: #333;
        background-color: #f0f2f6;
        font-family: Arial, sans-serif;
        margin: 0;
        padding: 0;
    }
    h1 {
        color: #333;
        text-align: center;
        margin-bottom: 20px;
    }
    .container {
        max-width: 800px;
        margin: auto;
        padding: 20px;
        background-color: #fff;
        border-radius: 8px;
        box-shadow: 0 0 10px rgba(0, 0, 0, 0.1);
    }
    .widget-label {
        color: #555;
    }
    .upload-section {
        margin-bottom: 20px;
    }
    .upload-container {
        display: flex;
        flex-wrap: wrap;
        gap: 10px;
    }
    .uploaded-image {
        width: 100px;
        height: auto;
        border-radius: 8px;
        overflow: hidden;
        box-shadow: 0 0 5px rgba(0, 0, 0, 0.1);
    }
    .download-btn {
        margin-top: 20px;
        text-align: center;
    }
    </style>
    """,
    unsafe_allow_html=True
)

# 앱 제목
st.title('테스트결과서 도우미')

# 플랫폼 선택
platforms = st.multiselect('플랫폼을 선택하세요 (복수 선택 가능)', ['iOS', 'AOS', 'HTS', 'MINTs', '홈페이지', '기타'])

# 업로드할 엑셀 파일
uploaded_excel = st.file_uploader("테스트결과서 엑셀 파일을 업로드하세요", type=['xlsx'])

# 엑셀 파일과 이미지 처리
if uploaded_excel and platforms:
    original_filename = uploaded_excel.name
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        tmp.write(uploaded_excel.getvalue())
        wb = load_workbook(tmp.name)

        for platform in platforms:
            ws = wb.create_sheet(title=f'{platform} 이미지 캡처')
            start_cell = 'A2'  # 기본 시작 셀 주소
            # 'A2'로부터 시작 행과 열 번호를 추출
            start_row, start_col = coordinate_to_tuple(start_cell)
            max_pixels = 500 if platform in ['iOS', 'AOS'] else 1000
            max_images_per_row = 6 if platform in ['iOS', 'AOS'] else 3

            images = st.file_uploader(f"{platform} 이미지 파일을 업로드하세요", accept_multiple_files=True, type=['png', 'jpg', 'jpeg'], key=f"files_{platform}")
            if images:
                # 행 별로 가장 큰 이미지 높이 저장
                max_heights_per_row = []
                current_row_images = []
                current_row_index = start_row
                x_offset = start_col

                for index, uploaded_file in enumerate(images):
                    # 이미지 처리
                    pil_image = PILImage.open(uploaded_file)
                    original_width, original_height = pil_image.size
                    ratio = min(max_pixels / original_width, max_pixels / original_height)
                    new_width, new_height = int(original_width * ratio), int(original_height * ratio)
                    pil_image = pil_image.resize((new_width, new_height))

                    # 현재 행에 대한 이미지들의 최대 높이 계산
                    if len(current_row_images) < max_images_per_row:
                        current_row_images.append((new_width, new_height))
                    else:
                        max_heights_per_row.append(max(img[1] for img in current_row_images))
                        current_row_images = [(new_width, new_height)]
                        x_offset = start_col  # x 위치 리셋
                        current_row_index += max_heights_per_row[-1] // 20 + 2  # 두 셀 높이만큼 간격 추가

                    # 이미지 저장 및 추가
                    with tempfile.NamedTemporaryFile(delete=False, suffix='.png') as tmp_img:
                        pil_image.save(tmp_img.name)
                        img = OpenpyxlImage(tmp_img.name)

                        # 이미지 추가
                        img.anchor = ws.cell(row=current_row_index, column=x_offset).coordinate
                        ws.add_image(img)

                        # 다음 이미지의 x 위치 계산
                        x_offset += new_width // 64 + 1

                # 마지막 행 처리
                if current_row_images:
                    max_heights_per_row.append(max(img[1] for img in current_row_images))

        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_excel:
            wb.save(tmp_excel.name)
            with open(tmp_excel.name, "rb") as f:
                st.download_button('수정된 엑셀 파일 다운로드', f, file_name=original_filename)
else:
    st.write('이미지 파일, 엑셀 파일 및 플랫폼을 선택해주세요.')








